import openpyxl

dataframe = openpyxl.load_workbook("produkty.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active

# Iterate the loop to read the cell values
for row in range(1, dataframe1.max_row + 1):
    # Assuming you have 4 columns (change the range accordingly)
    for col_num, col in enumerate(dataframe1.iter_cols(1, dataframe1.max_column), start=1):
        # Assign each column value to a variable
        variable_name = f"variable_{col_num}"
        print(variable_name)
        print(col[row - 1].value)
        # globals()[variable_name] = col[row - 1].value



    # Otwórz plik CSV i wczytaj zawartość
    # with open(csvfile, encoding='utf-8') as csvfile1:
    #     csv_reader = csv.reader(csvfile1, delimiter=',', quotechar='"', ensure_ascii=False)
    #     headers = next(csv_reader)  # wczytaj pierwszy wiersz, który zawiera nagłówki kolumn
    #
    #     # Sprawdź czy liczba kolumn w pliku CSV jest taka sama jak w słowniku kolumn
    #     print(headers)
    #
    #     for row in csv_reader:
    #         # Użyj formatowania stringów, aby uniknąć problemów z wyświetlaniem polskich znaków
    #         formatted_row = ", ".join(row)
    #         print(formatted_row)

        # if len(headers) != len(csv_file):
        #     self.errorLog.addWarning(f"File contains: {len(headers)} columns, expected: {len(csv_file)}")
        #     # print("Number of columns in CSV file does not match the number of objects in dictionary_file.")


# import_csv("przepisy.xlsx")