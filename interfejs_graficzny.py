import tkinter
from tkinter import *
from pathlib import Path
from tkinter.font import Font
from tkinter import messagebox, ttk
import webbrowser
import openpyxl

selected = []

def callback(url):
    webbrowser.open_new_tab(url)


def relative_to_assets1(path: str) -> Path:
    OUTPUT_PATH = Path(__file__).parent
    ASSETS_PATH = OUTPUT_PATH / Path(r"assets/frame0")
    return ASSETS_PATH / Path(path)


def relative_to_assets2(path: str) -> Path:
    OUTPUT_PATH = Path(__file__).parent
    ASSETS_PATH = OUTPUT_PATH / Path(r"assets/frame1")
    return ASSETS_PATH / Path(path)


def relative_to_assets3(path: str) -> Path:
    OUTPUT_PATH = Path(__file__).parent
    ASSETS_PATH = OUTPUT_PATH / Path(r"assets/frame2")
    return ASSETS_PATH / Path(path)


def relative_to_assets4(path: str) -> Path:
    OUTPUT_PATH = Path(__file__).parent
    ASSETS_PATH = OUTPUT_PATH / Path(r"assets/frame3")
    return ASSETS_PATH / Path(path)


def relative_to_assets5(path: str) -> Path:
    OUTPUT_PATH = Path(__file__).parent
    ASSETS_PATH = OUTPUT_PATH / Path(r"assets/frame4")
    return ASSETS_PATH / Path(path)


def relative_to_assets6(path: str) -> Path:
    OUTPUT_PATH = Path(__file__).parent
    ASSETS_PATH = OUTPUT_PATH / Path(r"assets/frame5")
    return ASSETS_PATH / Path(path)


def relative_to_assets7(path: str) -> Path:
    OUTPUT_PATH = Path(__file__).parent
    ASSETS_PATH = OUTPUT_PATH / Path(r"assets/frame6")
    return ASSETS_PATH / Path(path)


def relative_to_assets8(path: str) -> Path:
    OUTPUT_PATH = Path(__file__).parent
    ASSETS_PATH = OUTPUT_PATH / Path(r"assets/frame8")
    return ASSETS_PATH / Path(path)

def create_interface1(root):
    frame1 = Frame(root, bg="#F3F3F3")
    frame1.pack(fill='both', expand=True)

    canvas = Canvas(
        frame1,
        bg="#F3F3F3",
        height=600,
        width=1000,
        bd=0,
        highlightthickness=0,
        relief="ridge"
    )

    global image_image_1
    canvas.place(x=0, y=0)
    image_image_1 = PhotoImage(file=relative_to_assets1("image_1.png"))
    canvas.create_image(
        787.0,
        250.0,
        image=image_image_1
    )

    global button_image_1
    button_image_1 = PhotoImage(file=relative_to_assets1("button_1.png"))
    button_1 = Button(
        frame1,
        image=button_image_1,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame(interface2, interface1),
        relief="flat"
    )
    button_1.place(
        x=74.0,
        y=104.0,
        width=306.0,
        height=67.0
    )

    global button_image_2
    button_image_2 = PhotoImage(file=relative_to_assets1("button_2.png"))
    button_2 = Button(
        frame1,
        image=button_image_2,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame(interface2, interface1),
        relief="flat"
    )
    button_2.place(
        x=74.0,
        y=450.0,
        width=250.0,
        height=67.0
    )

    canvas.create_text(
        74.0,
        201.0,
        anchor="nw",
        text="Zdrowomix - Twoje Centrum Zdrowego Żywienia Rodziny! W tej innowacyjnej aplikacji znajdziesz wszystko, czego potrzebujesz do planowania żywienia dla swojej rodziny. Skutecznie, ekonomicznie i zero waste!",
        fill="#414141",
        font=("Inter Medium", 25 * -1),
        width="400"
    )

    return frame1


def create_interface2(root):
    frame2 = Frame(root, bg="#F3F3F3")
    frame2.pack(fill='both', expand=True)

    canvas = Canvas(
        frame2,
        bg="#F3F3F3",
        height=600,
        width=1000,
        bd=0,
        highlightthickness=0,
        relief="ridge"
    )

    canvas.place(x=0, y=0)
    canvas.create_text(
        18.0,
        17.0,
        anchor="nw",
        text="Zdrowomix",
        fill="#414141",
        font=("Inter Bold", 55 * -1)
    )
    global image_image_1_1
    image_image_1_1 = PhotoImage(
        file=relative_to_assets2("image_1.png"))
    image_1 = canvas.create_image(
        929.0,
        106.0,
        image=image_image_1_1
    )

    canvas.create_rectangle(
        33.0,
        173.0,
        251.6883544921875,
        515.4514465332031,
        fill="#F9F9F9",
        outline="")

    global button_image_1_1
    button_image_1_1 = PhotoImage(
        file=relative_to_assets2("button_1.png"))
    button_1 = Button(
        frame2,
        image=button_image_1_1,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame(interface3, interface2),
        relief="flat"
    )
    button_1.place(
        x=42.809814453125,
        y=464.618896484375,
        width=197.97975158691406,
        height=42.80643081665039
    )

    canvas.create_text(
        60.64581298828125,
        183.70166015625,
        anchor="nw",
        text="Kalkulator \nkalorii",
        fill="#414141",
        font=("Inter SemiBold", 25 * -1),
    )

    canvas.create_text(
        53.51141357421875,
        261.28826904296875,
        anchor="nw",
        text="Odkryj sekrety zdrowego żywienia! Nasz Kalkulator Kaloryczny dostosowuje się do Twoich indywidualnych potrzeb - wprowadź wiek, płeć oraz poziom aktywności, a otrzymasz spersonalizowane wyniki, aby utrzymać równowagę kaloryczną. ",
        fill="#414141",
        font=("Inter Medium", 16 * -1),
        width="200",
        justify="center"
    )

    canvas.create_rectangle(
        271.307861328125,
        173.0,
        489.9962158203125,
        515.4514465332031,
        fill="#F9F9F9",
        outline="")

    global button_image_2_1
    button_image_2_1 = PhotoImage(
        file=relative_to_assets2("button_2.png"))
    button_2 = Button(
        frame2,
        image=button_image_2_1,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame(interface5, interface2),  # tutaj
        relief="flat"
    )
    button_2.place(
        x=281.662353515625,
        y=464.618896484375,
        width=197.97975158691406,
        height=42.80643081665039
    )

    canvas.create_text(
        295.93115234375,
        183.70166015625,
        anchor="nw",
        text="Spis\nproduktów",
        fill="#414141",
        font=("Inter SemiBold", 25 * -1),
    )

    canvas.create_text(
        292.3638916015625,
        261.28826904296875,
        anchor="nw",
        text="Odkryj przepisy, dzięki Twoim ulubionym produktom! Nasz Spis Produktów ułatwia planowanie żywienia rodziny - znajdź lub dodaj ulubione składniki, a szybko znajdziesz idealny przepis dopasowany do dostępnych produktów w Twoim domu. ",
        fill="#414141",
        font=("Inter Medium", 16 * -1),
        width="200",
        justify="center"
    )

    canvas.create_rectangle(
        509.6158447265625,
        173.0,
        728.30419921875,
        515.4514465332031,
        fill="#F9F9F9",
        outline="")

    global button_image_3_1
    button_image_3_1 = PhotoImage(
        file=relative_to_assets2("button_3.png"))
    button_3 = Button(
        frame2,
        image=button_image_3_1,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame(interface6, interface2),
        relief="flat"
    )
    button_3.place(
        x=519.6229248046875,
        y=464.618896484375,
        width=197.97975158691406,
        height=42.80643081665039
    )

    canvas.create_text(
        535.67529296875,
        183.70166015625,
        anchor="nw",
        text="Porady\nżywieniowe",
        fill="#414141",
        font=("Inter SemiBold", 25 * -1),
    )

    canvas.create_text(
        531.21630859375,
        261.28826904296875,
        anchor="nw",
        text="Zdrowe Wybory, Szczęśliwa Rodzina! Odkryj porady żywieniowe na naszej stronie, wspomagającej planowanie posiłków dla całej rodziny. Znajdziesz tu praktyczne wskazówki dotyczące zrównoważonej diety, dostosowane do różnych preferencji żywieniowych!",
        fill="#414141",
        font=("Inter Medium", 16 * -1),
        width="200",
        justify="center"
    )

    canvas.create_rectangle(
        748.0,
        173.0,
        966.6883544921875,
        515.4514465332031,
        fill="#F9F9F9",
        outline="")

    global button_image_4_1
    button_image_4_1 = PhotoImage(
        file=relative_to_assets2("button_4.png"))
    button_4 = Button(
        frame2,
        image=button_image_4_1,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame(interface7, interface2),
        relief="flat"
    )
    button_4.place(
        x=758.007080078125,
        y=464.618896484375,
        width=197.97975158691406,
        height=42.80643081665039
    )

    canvas.create_text(
        774.0595703125,
        183.70166015625,
        anchor="nw",
        text="Sprawdź co\njesz!",
        fill="#414141",
        font=("Inter SemiBold", 25 * -1),
    )

    canvas.create_text(
        769.6005859375,
        261.28826904296875,
        anchor="nw",
        text="Odkryj skład!Na naszej stronie zaplanujesz zdrowe posiłki dla rodziny - wybierz produkt, a otrzymasz natychmiastowe informacje o kaloriach, białkach, tłuszczach i węglowodanach. Możesz błyskawicznie ocenić, czy produkt spełnia Twoje żywieniowe oczekiwania! ",
        fill="#414141",
        font=("Inter Medium", 16 * -1),
        width="200",
        justify="center"
    )

    return frame2


def create_interface3(root):
    frame3 = Frame(root, bg="#F3F3F3")
    frame3.pack(fill='both', expand=True)

    canvas = Canvas(
        frame3,
        bg="#F3F3F3",
        height=600,
        width=1000,
        bd=0,
        highlightthickness=0,
        relief="ridge"
    )

    global button_image_1_2
    canvas.place(x=0, y=0)
    button_image_1_2 = PhotoImage(
        file=relative_to_assets3("button_1.png"))
    button_1 = Button(
        frame3,
        image=button_image_1_2,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame(interface2, interface3),
        relief="flat"
    )
    button_1.place(
        x=18.0,
        y=17.0,
        width=306.0,
        height=67.0
    )
    global button_image_2_2
    button_image_2_2 = PhotoImage(
        file=relative_to_assets3("button_2.png"))
    button_2 = Button(
        frame3,
        image=button_image_2_2,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame(interface2, interface3),
        relief="flat"
    )
    button_2.place(
        x=829.0,
        y=-20.0,
        width=201.0,
        height=150.0
    )

    canvas.create_text(
        20.0,
        101.0,
        anchor="nw",
        text="Kalkulator kalorii \n",
        fill="#414141",
        font=("Inter SemiBold", 40 * -1)
    )

    canvas.create_text(
        18.0,
        185.0,
        anchor="nw",
        text="Kalkulator kalorii w naszej aplikacji wspomagajacej planowanie żywienia rodziny to klucz do zdrowego stylu życia dla Ciebie i Twojej rodziny. Dzięki niemu, uzyskasz dostęp do spersonalizowanych danych o zapotrzebowaniu kalorycznym, które uwzględniają Twoje unikalne cechy, takie jak wiek, płeć, poziom aktywności fizycznej oraz wagę. To narzędzie nie tylko ułatwi Ci osiągnięcie celów zdrowotnych i dietetycznych, ale również pomoże w planowaniu zbilansowanych posiłków dla całej rodziny!",
        fill="#414141",
        font=("Inter SemiBold", 20 * -1),
        width="400"
    )

    canvas.create_text(
        466.0,
        255.0,
        anchor="nw",
        text="Waga:",
        fill="#414141",
        font=("Inter Medium", 25 * -1)
    )

    global entry_image_1_2
    entry_image_1_2 = PhotoImage(
        file=relative_to_assets3("entry_1.png"))
    entry_bg_1 = canvas.create_image(
        667.0,
        313.0,
        image=entry_image_1_2
    )
    entry_1 = Entry(
        frame3,
        bd=0,
        bg="#F9F9F9",
        fg="#000716",
        highlightthickness=0
    )
    entry_1.place(
        x=465.0,
        y=294.0,
        width=404.0,
        height=36.0
    )

    canvas.create_text(
        466.0,
        442.0,
        anchor="nw",
        text="Płeć",
        fill="#414141",
        font=("Inter Medium", 25 * -1)
    )

    canvas.create_text(
        466.0,
        159.0,
        anchor="nw",
        text="Wiek:",
        fill="#414141",
        font=("Inter Medium", 25 * -1)
    )

    global entry_image_3_2
    entry_image_3_2 = PhotoImage(
        file=relative_to_assets3("entry_3.png"))
    entry_bg_3 = canvas.create_image(
        667.0,
        217.0,
        image=entry_image_3_2
    )
    entry_3 = Entry(
        frame3,
        bd=0,
        bg="#F9F9F9",
        fg="#000716",
        highlightthickness=0
    )
    entry_3.place(
        x=465.0,
        y=198.0,
        width=404.0,
        height=36.0
    )

    canvas.create_text(
        466.0,
        71.0,
        anchor="nw",
        text="Podaj imię:",
        fill="#414141",
        font=("Inter Medium", 25 * -1)
    )
    global entry_image_4_2
    entry_image_4_2 = PhotoImage(
        file=relative_to_assets3("entry_4.png"))
    entry_bg_4 = canvas.create_image(
        667.0,
        129.0,
        image=entry_image_4_2
    )
    entry_4 = Entry(
        frame3,
        bd=0,
        bg="#F9F9F9",
        fg="#000716",
        highlightthickness=0
    )
    entry_4.place(
        x=465.0,
        y=110.0,
        width=404.0,
        height=36.0
    )

    canvas.create_text(
        466.0,
        350.0,
        anchor="nw",
        text="Aktywność",
        fill="#414141",
        font=("Inter Medium", 25 * -1)
    )

    global button_image_3_2
    button_image_3_2 = PhotoImage(
        file=relative_to_assets3("button_3.png"))
    button_3 = Button(
        frame3,
        image=button_image_3_2,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame(interface4, interface3),
        relief="flat"
    )
    button_3.place(
        x=465.0,
        y=539.0,
        width=404.0,
        height=60.0
    )

    canvas.create_text(
        551.0,
        17.0,
        anchor="nw",
        text="Twoja dieta:",
        fill="#414141",
        font=("Inter SemiBold", 40 * -1)
    )

    aktywnosci = ['niska', 'średnia', 'wysoka']
    branch_label = tkinter.Label(frame3, text='')
    branch_label.place(x=473.0, y=400.0)

    branch_dropdown = ttk.Combobox(frame3, values=aktywnosci)
    branch_dropdown.place(x=473.0, y=400.0)  # Dostosuj y według potrzeb

    plec = ['Kobieta', 'Mężczyzna']
    branch_label = tkinter.Label(frame3, text='')
    branch_label.place(x=473.0, y=480.0)

    branch_dropdown = ttk.Combobox(frame3, values=plec)
    branch_dropdown.place(x=473.0, y=480.0)  # Dostosuj y według potrzeb

    return frame3


def create_interface4(root):
    frame4 = Frame(root, bg="#F3F3F3")
    frame4.pack(fill='both', expand=True)

    canvas = Canvas(
        frame4,
        bg="#F3F3F3",
        height=600,
        width=1000,
        bd=0,
        highlightthickness=0,
        relief="ridge"
    )

    canvas.place(x=0, y=0)

    global button_image_1_4
    button_image_1_4 = PhotoImage(
        file=relative_to_assets4("button_1.png"))
    button_1 = Button(
        frame4,
        image=button_image_1_4,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame(interface2, interface4),
        relief="flat"
    )
    button_1.place(
        x=18.0,
        y=17.0,
        width=306.0,
        height=67.0
    )

    global button_image_2_4
    button_image_2_4 = PhotoImage(
        file=relative_to_assets4("button_2.png"))
    button_2 = Button(
        frame4,
        image=button_image_2_4,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame(interface2, interface4),
        relief="flat"
    )
    button_2.place(
        x=829.0,
        y=18.0,
        width=201.0,
        height=177.0
    )

    canvas.create_text(
        20.0,
        101.0,
        anchor="nw",
        text="Kalkulator kalorii \n",
        fill="#414141",
        font=("Inter SemiBold", 40 * -1)
    )

    canvas.create_text(
        102.0,
        159.0,
        anchor="nw",
        text="Twoje zapotrzebowanie kaloryczne wynosi : 10928kcal",
        fill="#414141",
        font=("Inter Medium", 30 * -1)
    )

    canvas.create_text(
        102.0,
        224.0,
        anchor="nw",
        text="Twoje indywidualne zapotrzebowanie kaloryczne zostało obliczone! Na podstawie podanych danych – wieku, wagi, poziomu aktywności fizycznej oraz płci – Twoje dzienne zapotrzebowanie kaloryczne wynosi [wpisz wynik kalorii].",
        fill="#414141",
        font=("Inter Medium", 25 * -1),
        width="400"
    )

    global image_image_1_4
    image_image_1_4 = PhotoImage(
        file=relative_to_assets4("image_1.png"))
    image_1 = canvas.create_image(
        792.0,
        396.0,
        image=image_image_1_4
    )

    global button_image_3_4
    button_image_3_4 = PhotoImage(
        file=relative_to_assets4("button_3.png"))
    button_3 = Button(
        frame4,
        image=button_image_3_4,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame(interface2, interface4),
        relief="flat"
    )
    button_3.place(
        x=102.0,
        y=484.0,
        width=276.0,
        height=60.0
    )
    return frame4


def create_interface5(root):
    frame5 = Frame(root, bg="#F3F3F3")
    frame5.pack(fill='both', expand=True)

    canvas = Canvas(
        frame5,
        bg="#F3F3F3",
        height=600,
        width=1000,
        bd=0,
        highlightthickness=0,
        relief="ridge"
    )

    canvas.place(x=0, y=0)
    global button_image_1_5
    button_image_1_5 = PhotoImage(
        file=relative_to_assets5("button_1.png"))
    button_1 = Button(
        frame5,
        image=button_image_1_5,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame(interface2, interface5),
        relief="flat"
    )
    button_1.place(
        x=21.0078125,
        y=28.71484375,
        width=297.8369140625,
        height=42.87109375
    )
    global button_image_2_5
    button_image_2_5 = PhotoImage(
        file=relative_to_assets5("button_2.png"))
    button_2 = Button(
        frame5,
        image=button_image_2_5,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame(interface2, interface5),
        relief="flat"
    )
    button_2.place(
        x=829.0,
        y=18.0,
        width=201.0,
        height=177.0
    )

    global entry_image_1_6
    entry_image_1_6 = PhotoImage(
        file=relative_to_assets5("entry_1.png"))
    entry_bg_1 = canvas.create_image(
        578.5,
        51.0,
        image=entry_image_1_6
    )
    entry_1 = Entry(
        frame5,
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0
    )
    entry_1.place(
        x=500.0,
        y=35.0,
        width=157.0,
        height=30.0
    )
    global button_image_3_5
    button_image_3_5 = PhotoImage(
        file=relative_to_assets5("button_3.png"))
    button_3 = Button(
        frame5,
        image=button_image_3_5,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: print("button_3 clicked"), #search_for_product(),
        relief="flat"
    )
    button_3.place(
        x=672.0,
        y=32.0,
        width=157.0,
        height=37.0
    )
    global image_image_1_5
    image_image_1_5 = PhotoImage(
        file=relative_to_assets5("image_1.png"))
    image_1 = canvas.create_image(
        774.0,
        375.0,
        image=image_image_1_5
    )

    canvas.create_text(
        20.0,
        101.0,
        anchor="nw",
        text="Spis produktów ",
        fill="#414141",
        font=("Inter SemiBold", 40 * -1)
    )

    # products = {"Cytryna": ["dawd", "yviuj"], "Jabłko": ["adwd", "kjhbk"], "Cynamon": ["awd", "awd"],
    #             "Arbuz": ["awda", "ohih8oi"], "Melon": ["dawd", "wadad"], 'Cebula': ["wad", "wda"],
    #             'Sezam': ["wad", "awdwa"], 'Maliny': ["awd", "duabsuiok"]}

    dataframe = openpyxl.load_workbook("produkty.xlsx")
    dataframe1 = dataframe.active

    products = []

    for row in range(2, dataframe1.max_row + 1):
        # products.append(dataframe1.cell(row=row, column=2).value)
        products.append(dataframe1.cell(row=row, column=2).value)
        # products[dataframe1.cell(row=row, column=2).value] = []

    selected_items = {}

    def update_selection(event):
        # Pobierz aktualnie wybrany element
        current_selection = product_listbox5.curselection()
        for item in current_selection:
            if item in selected_items:
                # Jeśli element był już wybrany, odznacz go
                product_listbox5.itemconfig(item, {'fg': 'black'})
                del selected_items[item]
            else:
                # Zaznacz nowy element
                product_listbox5.itemconfig(item, {'fg': 'green'})
                selected_items[item] = True

    product_listbox5 = Listbox(frame5, font=Font(size=15), selectmode='multiple')
    scrollbar5 = Scrollbar(frame5, orient="vertical", command=product_listbox5.yview)
    product_listbox5.config(yscrollcommand=scrollbar5.set, selectforeground='#06982B', selectbackground='#F3F3F3')

    product_listbox5.place(x=28, y=185, width=250, height=350)
    scrollbar5.place(x=265, y=185, height=350)

    for product in products:
        product_listbox5.insert("end", product),

    global button_image_4_5
    button_image_4_5 = PhotoImage(
        file=relative_to_assets5("button_4.png"))
    button_4 = Button(
        frame5,
        image=button_image_4_5,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame_5_to_8(interface5, product_listbox5.curselection(), products),
        relief="flat"
    )
    button_4.place(
        x=367.0,
        y=551.0,
        width=184.0,
        height=46.0
    )

    return frame5


def create_interface6(root):
    def update_product_info6(event):
        try:
            selection6 = product_listbox6.get(product_listbox6.curselection())
            product_info6.config(text=f"{selection6}\n{products[selection6][0]}")
            link = products[selection6][1]
            link_label.config(text=link, fg="blue", cursor="hand2")
            link_label.bind("<Button-1>", lambda e: callback(link))
        except Exception as e:
            print(e)

    frame6 = Frame(root, bg="#F3F3F3")
    frame6 = Frame(root, bg="#F3F3F3")
    frame6.pack(fill='both', expand=True)

    canvas = Canvas(
        frame6,
        bg="#F3F3F3",
        height=600,
        width=1000,
        bd=0,
        highlightthickness=0,
        relief="ridge"
    )

    canvas.place(x=0, y=0)
    global button_image_1_6
    button_image_1_6 = PhotoImage(
        file=relative_to_assets8("button_1.png"))
    button_1 = Button(
        frame6,
        image=button_image_1_6,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame(interface2, interface6),
        relief="flat"
    )
    button_1.place(
        x=21.0078125,
        y=28.71484375,
        width=297.8369140625,
        height=42.87109375
    )
    global button_image_2_6
    button_image_2_6 = PhotoImage(
        file=relative_to_assets5("button_2.png"))
    button_2 = Button(
        frame6,
        image=button_image_2_6,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame(interface2, interface6),
        relief="flat"
    )
    button_2.place(
        x=829.0,
        y=18.0,
        width=201.0,
        height=177.0
    )
    global entry_image_1_6
    entry_image_1_6 = PhotoImage(
        file=relative_to_assets6("entry_1.png"))
    entry_bg_1 = canvas.create_image(
        578.5,
        51.0,
        image=entry_image_1_6
    )
    entry_1 = Entry(
        frame6,
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0
    )
    entry_1.place(
        x=500.0,
        y=35.0,
        width=157.0,
        height=30.0
    )
    global button_image_3_6
    button_image_3_6 = PhotoImage(
        file=relative_to_assets5("button_3.png"))
    button_3 = Button(
        frame6,
        image=button_image_3_6,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: print("button_3 clicked"),
        relief="flat"
    )
    button_3.place(
        x=672.0,
        y=32.0,
        width=157.0,
        height=37.0
    )
    global image_image_1_6
    image_image_1_6 = PhotoImage(
        file=relative_to_assets5("image_1.png"))
    image_1 = canvas.create_image(
        774.0,
        375.0,
        image=image_image_1_6
    )

    canvas.create_text(
        20.0,
        101.0,
        anchor="nw",
        text="Artykuły o tematyce zdrowego żywienia",
        fill="#414141",
        font=("Inter SemiBold", 20 * -1)
    )

    products = {"Jak jeść zdrowo i tanio": [
        "1. Kupuj sezonowe\nwarzywa i owoce!\nTo jest podstawa.\nW sezonie warzywa\ni owoce są najtańsze,\nale też najsmaczniejsze\ni mają najwięcej\nwitamin.\n\n""2. Korzystaj z mrożonek.\nŚwieże, surowe\nwarzywa i owoce są\noczywiście najlepsze(...).",
        "https://tiny.pl/cz8zb"], "Jak zacząć się zdrowo odżywiać": [
        "Jak mówią - nie od razu\nRzym zbudowano.\nTak samo jest ze zdrowym\njedzeniem. Z dnia na dzień\nzmiany nie przychodzą łatwo.\nEfekty długofalowe dobrze\nosiągać metodą małych kroków.\nW ten sposób dajemy sobie\nczas na przyzwyczajenie się\ndo zmian. Poniżej przedstawiamy\nlistę jak to uzyskać.\nZmiany wprowadzaj stopniowo(...).",
        "https://tiny.pl/cz83x"], "Jak wykształcić zdrowe nawyki\n żywieniowe u dziecka?": [
        "(...) Zdrowa żywność i właściwe\nnawyki żywieniowe są\npodstawą zdrowego życia\ni przynoszą korzyści\ndla osób w każdym wieku.\nUczenie dziecka zdrowego\nodżywiania od najmłodszych\nlat pomoże mu w budowaniu\npozytywnego stosunku\ndo jedzenia w dorosłym\nżyciu. Kształtowanie tych\nnawyków zaowocuje zdrowiem\noraz dobrą zabawą – nie\ntylko dla dziecka, ale i\ndla całej rodziny!(...)",
        "https://tiny.pl/cz83c"]
                }

    product_listbox6 = Listbox(frame6, font=Font(size=10))
    scrollbar6 = Scrollbar(frame6, orient="vertical", command=product_listbox6.yview)
    product_listbox6.config(yscrollcommand=scrollbar6.set, selectforeground='#06982B', selectbackground='#F3F3F3')

    product_listbox6.place(x=28, y=185, width=250, height=350)
    scrollbar6.place(x=215, y=185, height=350)

    for product in products.keys():
        product_listbox6.insert("end", product)  # tutaj

        product_listbox6 = Listbox(frame6, font=Font(size=15))
        scrollbar6 = Scrollbar(frame6, orient="vertical", command=product_listbox6.yview)
        product_listbox6.config(yscrollcommand=scrollbar6.set, selectforeground='#06982B', selectbackground='#F3F3F3')

        product_listbox6.place(x=28, y=185, width=190, height=350)
        scrollbar6.place(x=215, y=185, height=350)

        for product in products.keys():
            product_listbox6.insert("end", product)

        product_listbox6 = Listbox(frame6, font=Font(size=15))
        scrollbar6 = Scrollbar(frame6, orient="vertical", command=product_listbox6.yview)
        product_listbox6.config(yscrollcommand=scrollbar6.set, selectforeground='#06982B', selectbackground='#F3F3F3', )

        product_listbox6.place(x=28, y=185, width=190, height=350)
        scrollbar6.place(x=215, y=185, height=350)

        for product in products.keys():
            product_listbox6.insert("end", product)

        product_info6 = Label(frame6, text="", font=("Inter SemiBold", 15))
        product_info6.place(x=230, y=185, width=300, height=400)

        product_listbox6.bind("<<ListboxSelect>>", update_product_info6)
        product_listbox6.selection_set(0)

        canvas.create_text(
            505.0,
            68.0,
            anchor="nw",
            text="Artykuł:",
            fill="#414141",
            font=("Inter Medium", 35 * -1)
        )

        Artykuł = canvas.create_text(
            505.0,
            118.0,
            anchor="nw",
            text="",
            fill="#414141",
            font=("Inter Medium", 25 * -1),
        )
        link_label = Label(frame6, text="", font=("Inter Medium", 25), fg="blue", cursor="hand2")
        link_label.place(x=505.0, y=105.0)  # Możesz dostosować położenie według potrzeb

    return frame6


def create_interface7(root):
    def update_product_info7(event):
        try:
            selection = product_listbox7.get(product_listbox7.curselection())
            # Aktualizacja informacji o produkcie
            # Tutaj możesz zastąpić to rzeczywistymi danymi dla każdego produktu
            product_info7.config(
                text=f"kcal {products[selection][0]}\nbiałka {products[selection][1]}g\ntłuszcze {products[selection][2]}g\nwęglowodany {products[selection][3]}g")
        except:
            pass

    frame7 = Frame(root, bg="#F3F3F3")
    frame7.pack(fill='both', expand=True)

    canvas = Canvas(
        frame7,
        bg="#F3F3F3",
        height=600,
        width=1000,
        bd=0,
        highlightthickness=0,
        relief="ridge"
    )

    canvas.place(x=0, y=0)

    global button_image_1_7
    button_image_1_7 = PhotoImage(
        file=relative_to_assets7("button_1.png"))
    button_1 = Button(
        frame7,
        image=button_image_1_7,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame(interface2, interface7),
        relief="flat"
    )
    button_1.place(
        x=18.0,
        y=17.0,
        width=306.0,
        height=67.0
    )

    global button_image_2_7
    button_image_2_7 = PhotoImage(
        file=relative_to_assets7("button_2.png"))
    button_2 = Button(
        frame7,
        image=button_image_2_7,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame(interface2, interface7),
        relief="flat"
    )
    button_2.place(
        x=829.0,
        y=-20,
        width=201.0,
        height=177.0
    )

    global entry_image_1_7
    entry_image_1_7 = PhotoImage(
        file=relative_to_assets7("entry_1.png"))
    entry_bg_1 = canvas.create_image(
        578.5,
        51.0,
        image=entry_image_1_7
    )
    entry_1 = Entry(
        frame7,
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0
    )
    entry_1.place(
        x=500.0,
        y=35.0,
        width=157.0,
        height=30.0
    )

    canvas.create_text(
        536.0,
        31.0,
        anchor="nw",
        text="wpisz",
        fill="#414141",
        font=("Inter SemiBold", 30 * -1)
    )

    global button_image_3_7
    button_image_3_7 = PhotoImage(
        file=relative_to_assets7("button_3.png"))
    button_3 = Button(
        frame7,
        image=button_image_3_7,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: print("oandosadndd"),
        relief="flat"
    )
    button_3.place(
        x=672.0,
        y=31.0,
        width=157.0,
        height=38.0
    )

    canvas.create_text(
        20.0,
        101.0,
        anchor="nw",
        text="Sprawdź co jesz!",
        fill="#414141",
        font=("Inter SemiBold", 40 * -1)
    )

    global image_image_1_7
    image_image_1_7 = PhotoImage(
        file=relative_to_assets7("image_1.png"))
    image_1 = canvas.create_image(
        674.0,
        370.0,
        image=image_image_1_7
    )
    # products = {"Cytryna": ["kcal: \n białka: \n tłuszcze: \n, węglowodany: \n", "yviuj"], "Jabłko": ["kcal: \n białka: \n tłuszcze: \n, węglowodany: \n", "kjhbk"], "Cynamon": ["kcal: \n białka: \n tłuszcze: \n, węglowodany: \n", "awd"],
    #             "Arbuz": ["kcal: \n białka: \n tłuszcze: \n, węglowodany: \n", "ohih8oi"], "Melon": ["kcal: \n białka: \n tłuszcze: \n, węglowodany: \n", "wadad"], 'Cebula': ["kcal: \n białka: \n tłuszcze: \n, węglowodany: \n" , "wda"],
    #             'Mleko spożywcze 2%': ["kcal: \n białka: \n tłuszcze: \n, węglowodany: \n", "awdwa"], 'Bulion warzywny': ["kcal: \n białka: \n tłuszcze: \n, węglowodany: \n", "duabsuiok"]}

    dataframe = openpyxl.load_workbook("produkty.xlsx")
    dataframe1 = dataframe.active

    products = {}

    for row in range(2, dataframe1.max_row + 1):
        # products.append(dataframe1.cell(row=row, column=2).value)
        products[dataframe1.cell(row=row, column=2).value] = [dataframe1.cell(row=row, column=3).value,
                                                              dataframe1.cell(row=row, column=4).value,
                                                              dataframe1.cell(row=row, column=5).value,
                                                              dataframe1.cell(row=row, column=6).value]

    product_listbox7 = Listbox(frame7, font=Font(size=15))
    scrollbar7 = Scrollbar(frame7, orient="vertical", command=product_listbox7.yview)
    product_listbox7.config(yscrollcommand=scrollbar7.set, selectforeground='#06982B', selectbackground='#F3F3F3')

    product_listbox7.place(x=28, y=185, width=190, height=350)
    scrollbar7.place(x=215, y=185, height=350)

    for product in products.keys():
        product_listbox7.insert("end", product)

    product_info7 = Label(frame7, text="", font=("Inter SemiBold", 15), bg="#F3F3F3")
    product_info7.place(x=270, y=185, width=200, height=200)

    product_listbox7.bind("<<ListboxSelect>>", update_product_info7)
    product_listbox7.selection_set(0)

    return frame7


# nowy frame porady żywnościowe
def create_interface8(root):
    def update_product_info8(event):
        try:
            selection8 = product_listbox6.get(product_listbox6.curselection())
            # Aktualizacja informacji o produkcie
            # Tutaj możesz zastąpić to rzeczywistymi danymi dla każdego produktu
            i = 0
            for i in range(len(products)):
                if products[i][0] == selection8:
                    product_info6.config(text=f"{selection8}\n{products[i][1]}")
                    link = products[i][2]
                    link_label.config(text=link, fg= "blue", cursor="hand2")
                    link_label.bind("<Button-1>", lambda e: callback(link))
                    break
        except Exception as e:
            print(e)

    frame6 = Frame(root, bg="#F3F3F3")
    frame8 = Frame(root, bg="#F3F3F3")
    frame8.pack(fill='both', expand=True)

    canvas = Canvas(
        frame8,
        bg="#F3F3F3",
        height=600,
        width=1000,
        bd=0,
        highlightthickness=0,
        relief="ridge"
    )

    canvas.place(x=0, y=0)
    global button_image_1_8
    button_image_1_8 = PhotoImage(
        file=relative_to_assets8("button_1.png"))
    button_1 = Button(
        frame8,
        image=button_image_1_8,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame(interface2, interface8),
        relief="flat"
    )
    button_1.place(
        x=21.0078125,
        y=28.71484375,
        width=297.8369140625,
        height=42.87109375
    )
    global button_image_2_8
    button_image_2_8 = PhotoImage(
        file=relative_to_assets5("button_2.png"))
    button_2 = Button(
        frame8,
        image=button_image_2_8,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: switch_frame(interface2, interface8),
        relief="flat"
    )
    button_2.place(
        x=829.0,
        y=18.0,
        width=201.0,
        height=177.0
    )
    global entry_image_1_8
    entry_image_1_8 = PhotoImage(
        file=relative_to_assets8("entry_1.png"))
    entry_bg_1 = canvas.create_image(
        578.5,
        51.0,
        image=entry_image_1_8
    )
    entry_1 = Entry(
        frame8,
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0
    )
    entry_1.place(
        x=500.0,
        y=35.0,
        width=157.0,
        height=30.0
    )
    global button_image_3_8
    button_image_3_8 = PhotoImage(
        file=relative_to_assets5("button_3.png"))
    button_3 = Button(
        frame8,
        image=button_image_3_8,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: print("button_3 clicked"),
        relief="flat"
    )
    button_3.place(
        x=672.0,
        y=32.0,
        width=157.0,
        height=37.0
    )
    global image_image_1_8
    image_image_1_8 = PhotoImage(
        file=relative_to_assets8("image_1.png"))
    image_1 = canvas.create_image(
        774.0,
        375.0,
        image=image_image_1_8
    )

    canvas.create_text(
        20.0,
        101.0,
        anchor="nw",
        text="Zbiór przepisów",
        fill="#414141",
        font=("Inter SemiBold", 40 * -1)
    )

    # products = {"Zupa marchewkowa\n z camembertem": ["Kremowa i aromatyczna zupa,\n w której główną rolę odgrywa\n słodka marchewka połączona\n z wykwintnym smakiem\n sera camembert.\n Idealna na ciepły, kojący posiłek,\n łącząca klasyczne smaki\n z nutą elegancji.\n", "https://tiny.pl/czs4h"], "Zupa krem\n z zielonego groszku": ["Delikatna i aksamitna\n zupa krem, która w pełni\n wydobywa świeży smak\n zielonego groszku.\n To idealne połączenie\n lekkości i świeżości,\n doskonałe na wiosenny obiad\n lub jako elegancka przystawka.\n", "https://tiny.pl/czs4q"], "Potrawka z kurczaka": ["Serwowana z warzywami,\n ta potrawka to kwintesencja\n domowego, pożywnego obiadu.\n Kawałki kurczaka duszone\n w delikatnym sosie z dodatkiem\n marchwi i groszku,\n idealne z ryżem lub kaszą.\n", "https://tiny.pl/czs4m"],
    #             "Zupa krem z cukinii": ["Łagodna i kremowa zupa,\n w której cukinia spotyka się\n z delikatnymi przyprawami,\n tworząc idealne danie\n na każdą porę roku.\n Prosta w przygotowaniu,\n ale pełna smaku, zupa\n ta to doskonały sposób\n na lekki obiad lub kolację.\n", "https://tiny.pl/czs4t"], "Sałatka z burakiem,\n fetą i malinami": ["Połączenie pieczonego buraka,\n słonego sera feta i\n słodkich malin tworzy\n niepowtarzalną sałatkę pełną\n kontrastów smakowych. Jest to\n idealna propozycja na lekki\n posiłek lub wykwintną przystawkę,\n która zachwyci swoją\n barwnością i smakiem.\n", "https://tiny.pl/czs47"], "Kakaowe naleśniki z masłem\n orzechowym i dżemem\n truskawkowym\n": ["Te słodkie i puszyste naleśniki\n kakaowe, podane z kremowym\n masłem orzechowym i domowym\n dżemem truskawkowym,\n to idealne śniadanie na słodko.\n", "https://tiny.pl/czs4w"],
    #             "Sałatka z buraka,\n koziego sera, czerwonej\n cebuli, kiełków i rukoli\n": ["Ta elegancka sałatka\n łączy pieczony burak\n z delikatnym kozim serem\n i pikantną czerwoną cebulą,\n tworząc wyrafinowaną\n kompozycję smaków.\n ", "https://tiny.pl/czs4d"], "Kanapka z polędwiczką,\n sałatą oraz ketchupem\n": ["Soczysta polędwiczka w\n połączeniu z chrupiącą sałatą\n i domowym keczupem tworzy\n idealną kanapkę na szybki\n i smaczny posiłek.\n", "https://tiny.pl/czs4f"], "Cebulaki\n": ["Chrupiące i pełne smaku\n, te cebulaki to idealna\n przekąska dla miłośników\n ostrego smaku, podkreślona\n delikatnym sosem i\n świeżymi ziołami.\n", " https://tiny.pl/czs41"], "Makaron z bocbem,\n boczkiem i parmezanem\n": ["Ten makaron łączy świeży\n smak bobu z chrupiącym boczkiem\n i wyrazistym parmezanem, tworząc\n danie pełne smaku i tekstury.\n ", "https://tiny.pl/czs4j"],
    #             "Koktajl\n truskawkowo-malinowy": ["Ten orzeźwiający napój łączy\n świeżość owoców z lekkością\n i prostotą przygotowania,\n będąc idealnym wyborem\n na upalne dni.\n Jego świeży smak i\n aromat pobudzą Twoje zmysły.\n", "https://tiny.pl/czs4p"], "Napój\n melonowo-truskawkowy": ["Ten orzeźwiający napój\n łączy świeżość owoców\n z lekkością i prostotą\n przygotowania, będąc idealnym\n wyborem na upalne dni.\n Jego świeży smak i\n aromat pobudzą Twoje zmysły.\n", "https://tiny.pl/czs4l"], "Sałatka z roszponki,\n mozzarelli i pomidorków\n": ["Ta kolorowa i pełna\n świeżości sałatka to\n doskonałe połączenie chrupiących\n warzyw i intensywnych dodatków,\n idealna jako lekki posiłek\n lub wykwintna przystawka.\n Jest nie tylko smaczna,\n ale i estetycznie podana.\n", " https://tiny.pl/czs44"], "Sałatka z brokuła\n i sera fety z\n sosem czosnkowym\n": ["Ta kolorowa sałatka\n łączy różnorodne tekstury i\n smaki, tworząc lekki i odżywczy\n posiłek, idealny na\n każdą porę roku.\n", "https://tiny.pl/czs42"],
    #             "Czekoladowe mini serniczki\n na zimno z malinami\n": ["To danie to połączenie\n tradycyjnych smaków z\n nowoczesną prezentacją,\n idealne dla tych, którzy\n szukają czegoś wyjątkowego.\n", "https://tiny.pl/czs4s"], "Makaron z truskawkami\n": ["Nietypowe połączenie makaronu\n z słodkimi truskawkami tworzy\n zaskakujące, ale harmonijne danie,\n idealne na letni obiad.\n", "https://tiny.pl/czs46"],"Kokosowa kasza manna\n z owocami\n": [" Egzotyczna kasza manna\n z dodatkiem kokosa i\n kolorowych owoców to\n zdrowy sposób na rozpoczęcie\n dnia pełnego energii.\n", "https://cutt.ly/FwZQ0rgG"], "Placki z kiełbasą,\n serem i warzywami\n": ["Te smażone placki,\n wypełnione kawałkami kiełbasy,\n topionym serem i świeżymi\n warzywami, to idealny wybór'n na sycący posiłek.\n", "https://cutt.ly/ZwZQ0ovH"],
    #             "Szarlotka\n": ["Klasyczna szarlotka,\n z delikatnym ciastem i słodkim\n nadzieniem jabłkowym, to\n idealny deser dla miłośników\n owocowych słodyczy.\n", "https://cutt.ly/iwZQ0dgB"], "Ryż zapiekany z jabłkami\n": [" Ten zapiekany ryż z\n kawałkami jabłek i delikatną\n nutą cynamonu to ciepły\n i pocieszający deser,\n doskonały na chłodniejsze dni\n", "https://cutt.ly/AwZQ0bGM"],"Zapiekanka makarona\n z brokułem i szynką\n": ["Ta sycąca zapiekanka\n łączy w sobie bogate\n smaki i tekstury, będąc\n  doskonałym rozwiązaniem\n  na rodzinny obiad.\n Jej warstwowa kompozycja\n zapewnia pełnię smaku\n przy każdym kęsie.\n", "https://cutt.ly/qwZQ0Ert"], "Zapiekanka makaronowa z fasolką,\n brokułem i wędzonym łososiem\n": ["Komfortowa i sycąca, ta\n  zapiekanka to idealne połączenie\n makaronu z warzywami i mięsem,\n zapieczone pod chrupiącą\n warstwą sera.\n Idealna na rodzinny obiad.\n", "https://cutt.ly/bwZQ0USs"],
    #             "Sałatka meksykańska": ["Ta lekka i odświeżająca\n sałatka to idealny wybór\n na zdrowy posiłek.\n Łączy świeże składniki\n z wyrafinowanymi dodatkami,\n tworząc doskonałą\n harmonię smaków.\n", "https://cutt.ly/6wZQ00OV"], "Sałatka z makaronu,\n tuńczyka i kukurydzy\n": ["Ta sałatka to idealne połączenie\n makaronu al dente, delikatnego\n tuńczyka i słodkiej kukurydzy,\n tworząc sycący, ale lekki\n posiłek. Świeża i pełna\n smaku, jest doskonała zarówno\n jako samodzielne danie, jak\n i wykwintny dodatek do obiadu.\n", "https://cutt.ly/zwZQ048c"],"Smoothie z płatkami jaglanymi,\n bananem, szpinakiem i kiwi\n": ["Orzeźwiające i pełne\n witamin smoothie to idealny\n sposób na rozpoczęcie dnia.\n Jego bogaty smak i\n kremowa konsystencja sprawią,\n że stanie się Twoim\n ulubionym napojem.\n", "https://cutt.ly/gwZQ2elB"], "Płatki jęczmienne z\n nasionami Chia i owocami\n": ["Te odżywcze płatki jęczmienne,\n wzbogacone nasionami chia\n i świeżymi owocami, to\n idealny start w dzień\n dla każdego, kto szuka zdrowego\n i sycącego śniadania. Ich\n bogaty w składniki odżywcze\n profil i słodki smak owoców\n zapewniają energię\n na cały poranek.\n", " https://cutt.ly/6wZQ2fHu"],
    #             "Serek wiejski z winogronami,\n orzechami włoskimi i miodem\n": ["Ta lekka i zdrowa przekąska\n łączy serek wiejski z\n słodyczą winogron,\n chrupkością orzechów i\n miodem, tworząc idealną\n kompozycję smaków.\n", "https://cutt.ly/SwZQ2xdK"], "Klopsiki z indyka z\n ziemniakami i zielonym groszkiem\n": ["Soczyste klopsiki z\n indyka podane z kremowym\n sosem, młodymi ziemniakami\n i groszkiem to idealna propozycja\n na sycący i smaczny obiad.\n", "https://cutt.ly/JwZQ2nfj"],"Smoothie z mango,\n arbuza i truskawek\n": ["Orzeźwiające i pełne witamin,\n to smoothie to idealny sposób\n na zdrowy i szybki posiłek.\n Jego kremowa konsystencja\n i bogaty smak owoców\n zadowolą każdego.\n", "https://cutt.ly/iwZQ2PFx"], "Gofry": ["To danie to doskonałe\n połączenie tradycji z \n nowoczesnymi akcentami, idealne\n dla poszukiwaczy kulinarnych wrażeń.\n ", "https://cutt.ly/pwZQ2GLO"],
    #             "Nocna owsianka z owocami": ["To danie to perfekcyjne\n połączenie smaków i aromatów,\n które zaskoczą i zachwycą\n każdego smakosza.\n Idealne dla tych,\n którzy szukają nowych\n kulinarnych wrażeń.\n", "https://cutt.ly/WwZQ2Xkv"], "Paluszki rybne z frytkami\n i surówką z czerwonej kapusty\n": [" Klasyczne danie w\n nowoczesnej odsłonie,\n idealne dla dzieci i dorosłych.\n Chrupiące paluszki rybne\n z frytkami i świeżą surówką\n to gwarancja smacznego posiłku.\n", "https://cutt.ly/rwZQ21Pt"],"Chlebek bananowy": ["Słodka i aromatyczna\n przekąska, idealna\n do kawy lub herbaty.\n Ten domowy chlebek,\n pełen smaku bananów,\n zadowoli każdego łasucha.\n", "https://cutt.ly/pwZQ23lb"], "Kotlety drobiowe z\n młodymi ziemniakami\n i sałatą\n": ["Soczyste kotlety drobiowe\n podane z młodymi ziemniakami\n i świeżą sałatą to idealny\n obiad dla całej rodziny.\n Proste, ale pełne smaku.\n", "https://cutt.ly/SwZQ26om"],"Chlebek czosnkowy": ["Aromatyczna przekąska,\n ten domowy chlebek,\n pełen smaku czosnku,\n zadowoli każdego łasucha.\n", "https://cutt.ly/cwZQ9yP8"], "Mocno czekoladowe trufle\n z mleka zagęszczonego i kakao\n": ["To danie to perfekcyjne\n połączenie smaków i aromatów,\n które zaskoczą i zachwycą\n każdego smakosza.\n Idealne dla tych,\n którzy szukają nowych\n kulinarnych wrażeń.\n", " https://cutt.ly/swZQ4PUs"],"Zupa fasolowa": ["Tradycyjna, sycąca zupa\n fasolowa, pełna aromatycznych\n przypraw i warzyw.\n Doskonała na chłodne dni,\n by rozgrzać ciało i duszę.\n", " https://cutt.ly/cwZQ9yP8"], "Tarta z burakiem,\n cebulą i kozim serem\n": [" Elegancka i pełna smaku\n tarta, łącząca słodycz buraka\n z wyrazistym smakiem koziego sera.\n Idealna na wykwintną kolację\n lub imprezę.\n", "https://cutt.ly/XwZQ9sez"], "Drożdzowe rogaliki\n z czekoladą\n":["Małe, słodkie rogaliki\n z czekoladą to idealna\n przekąska na każdą porę dnia.\n Ich miękka, drożdżowa tekstura\n i czekoladowe nadzienie sprawią,\n że nie będziesz mógł się oprzeć.\n","https://cutt.ly/OwZQ9g02"],"Ciasto francuskie z serem\n brie i żurawiną\n":["Wykwintne i eleganckie\n ciasto francuskie z serem brie\n i żurawiną to idealny wybór\n na wyjątkowe okazje.\n Jego delikatna tekstura i\n wyrafinowany smak zadowolą\n najbardziej wymagających gości.\n", "https://cutt.ly/HwZQ9zSo"]}
    products = [
        ["Zupa marchewkowa\n z camembertem",
         "Kremowa i aromatyczna zupa,\n w której główną rolę odgrywa\n słodka marchewka połączona\n z wykwintnym smakiem\n sera camembert.\n Idealna na ciepły, kojący posiłek,\n łącząca klasyczne smaki\n z nutą elegancji.\n",
         "https://tiny.pl/czs4h"],
        ["Zupa krem\n z zielonego groszku",
         "Delikatna i aksamitna\n zupa krem, która w pełni\n wydobywa świeży smak\n zielonego groszku.\n To idealne połączenie\n lekkości i świeżości,\n doskonałe na wiosenny obiad\n lub jako elegancka przystawka.\n",
         "https://tiny.pl/czs4q"],
        ["Potrawka z kurczaka",
         "Serwowana z warzywami,\n ta potrawka to kwintesencja\n domowego, pożywnego obiadu.\n Kawałki kurczaka duszone\n w delikatnym sosie z dodatkiem\n marchwi i groszku,\n idealne z ryżem lub kaszą.\n",
         "https://tiny.pl/czs4m"],
        ["Zupa krem z cukinii",
         "Łagodna i kremowa zupa,\n w której cukinia spotyka się\n z delikatnymi przyprawami,\n tworząc idealne danie\n na każdą porę roku.\n Prosta w przygotowaniu,\n ale pełna smaku, zupa\n ta to doskonały sposób\n na lekki obiad lub kolację.\n",
         "https://tiny.pl/czs4t"],
        ["Sałatka z burakiem,\n fetą i malinami",
         "Połączenie pieczonego buraka,\n słonego sera feta i\n słodkich malin tworzy\n niepowtarzalną sałatkę pełną\n kontrastów smakowych. Jest to\n idealna propozycja na lekki\n posiłek lub wykwintną przystawkę,\n która zachwyci swoją\n barwnością i smakiem.\n",
         "https://tiny.pl/czs47"],
        ["Kakaowe naleśniki z masłem\n orzechowym i dżemem\n truskawkowym\n",
         "Te słodkie i puszyste naleśniki\n kakaowe, podane z kremowym\n masłem orzechowym i domowym\n dżemem truskawkowym,\n to idealne śniadanie na słodko.\n",
         "https://tiny.pl/czs4w"],
        ["Sałatka z buraka,\n koziego sera, czerwonej\n cebuli, kiełków i rukoli\n",
         "Ta elegancka sałatka\n łączy pieczony burak\n z delikatnym kozim serem\n i pikantną czerwoną cebulą,\n tworząc wyrafinowaną\n kompozycję smaków.\n",
         "https://tiny.pl/czs4d"],
        ["Kanapka z polędwiczką,\n sałatą oraz ketchupem\n",
         "Soczysta polędwiczka w\n połączeniu z chrupiącą sałatą\n i domowym keczupem tworzy\n idealną kanapkę na szybki\n i smaczny posiłek.\n",
         "https://tiny.pl/czs4f"],
        ["Cebulaki",
         "Chrupiące i pełne smaku\n, te cebulaki to idealna\n przekąska dla miłośników\n ostrego smaku, podkreślona\n delikatnym sosem i\n świeżymi ziołami.\n",
         " https://tiny.pl/czs41"],
        ["Makaron z boczkiem,\n boczkiem i parmezanem\n",
         "Ten makaron łączy świeży\n smak bobu z chrupiącym boczkiem\n i wyrazistym parmezanem, tworząc\n danie pełne smaku i tekstury.\n ",
         "https://tiny.pl/czs4j"],
        ["Koktajl\n truskawkowo-malinowy",
         "Ten orzeźwiający napój łączy\n świeżość owoców z lekkością\n i prostotą przygotowania,\n będąc idealnym wyborem\n na upalne dni.\n Jego świeży smak i\n aromat pobudzą Twoje zmysły.\n",
         "https://tiny.pl/czs4p"],
        ["Napój\n melonowo-truskawkowy",
         "Ten orzeźwiający napój\n łączy świeżość owoców\n z lekkością i prostotą\n przygotowania, będąc idealnym\n wyborem na upalne dni.\n Jego świeży smak i\n aromat pobudzą Twoje zmysły.\n",
         "https://tiny.pl/czs4l"],
        ["Sałatka z roszponki,\n mozzarelli i pomidorków\n",
         "Ta kolorowa i pełna\n świeżości sałatka to\n doskonałe połączenie chrupiących\n warzyw i intensywnych dodatków,\n idealna jako lekki posiłek\n lub wykwintna przystawka.\n Jest nie tylko smaczna,\n ale i estetycznie podana.\n",
         " https://tiny.pl/czs44"],
        ["Sałatka z brokuła\n i sera fety z\n sosem czosnkowym\n",
         "Ta kolorowa sałatka\n łączy różnorodne tekstury i\n smaki, tworząc lekki i odżywczy\n posiłek, idealny na\n każdą porę roku.\n",
         "https://tiny.pl/czs42"],
        ["Czekoladowe mini serniczki\n na zimno z malinami\n",
         "To danie to połączenie\n tradycyjnych smaków z\n nowoczesną prezentacją,\n idealne dla tych, którzy\n szukają czegoś wyjątkowego.\n",
         "https://tiny.pl/czs4s"],
        ["Makaron z truskawkami\n",
         "Nietypowe połączenie makaronu\n z słodkimi truskawkami tworzy\n zaskakujące, ale harmonijne danie,\n idealne na letni obiad.\n",
         "https://tiny.pl/czs46"],
        ["Kokosowa kasza manna\n z owocami\n",
         " Egzotyczna kasza manna\n z dodatkiem kokosa i\n kolorowych owoców to\n zdrowy sposób na rozpoczęcie\n dnia pełnego energii.\n",
         "https://cutt.ly/FwZQ0rgG"],
        ["Placki z kiełbasą,\n serem i warzywami\n",
         "Te smażone placki,\n wypełnione kawałkami kiełbasy,\n topionym serem i świeżymi\n warzywami, to idealny wybór na sycący posiłek.\n",
         "https://cutt.ly/ZwZQ0ovH"],
        ["Szarlotka\n",
         "Klasyczna szarlotka,\n z delikatnym ciastem i słodkim\n nadzieniem jabłkowym, to\n idealny deser dla miłośników\n owocowych słodyczy.\n",
         "https://cutt.ly/iwZQ0dgB"],
        ["Ryż zapiekany z jabłkami\n",
         " Ten zapiekany ryż z\n kawałkami jabłek i delikatną\n nutą cynamonu to ciepły\n i pocieszający deser,\n doskonały na chłodniejsze dni\n",
         "https://cutt.ly/AwZQ0bGM"],
        ["Zapiekanka makarona\n z brokułem i szynką\n",
         "Ta sycąca zapiekanka\n łączy w sobie bogate\n smaki i tekstury, będąc\n  doskonałym rozwiązaniem\n  na rodzinny obiad.\n Jej warstwowa kompozycja\n zapewnia pełnię smaku\n przy każdym kęsie.\n",
         "https://cutt.ly/qwZQ0Ert"],
        ["Zapiekanka makaronowa z fasolką,\n brokułem i wędzonym łososiem\n",
         "Komfortowa i sycąca, ta\n  zapiekanka to idealne połączenie\n makaronu z warzywami i mięsem,\n zapieczone pod chrupiącą\n warstwą sera.\n Idealna na rodzinny obiad.\n",
         "https://cutt.ly/bwZQ0USs"],
        ["Sałatka meksykańska",
         "Ta lekka i odświeżająca\n sałatka to idealny wybór\n na zdrowy posiłek.\n Łączy świeże składniki\n z wyrafinowanymi dodatkami,\n tworząc doskonałą\n harmonię smaków.\n",
         "https://cutt.ly/6wZQ00OV"],
        ["Sałatka z makaronu,\n tuńczyka i kukurydzy\n",
         "Ta sałatka to idealne połączenie\n makaronu al dente, delikatnego\n tuńczyka i słodkiej kukurydzy,\n tworząc sycący, ale lekki\n posiłek. Świeża i pełna\n smaku, jest doskonała zarówno\n jako samodzielne danie, jak\n i wykwintny dodatek do obiadu.\n",
         "https://cutt.ly/zwZQ048c"],
        ["Smoothie z płatkami jaglanymi,\n bananem, szpinakiem i kiwi\n",
         "Orzeźwiające i pełne\n witamin smoothie to idealny\n sposób na rozpoczęcie dnia.\n Jego bogaty smak i\n kremowa konsystencja sprawią,\n że stanie się Twoim\n ulubionym napojem.\n",
         "https://cutt.ly/gwZQ2elB"],
        ["Płatki jęczmienne z\n nasionami Chia i owocami\n",
         "Te odżywcze płatki jęczmienne,\n wzbogacone nasionami chia\n i świeżymi owocami, to\n idealny start w dzień\n dla każdego, kto szuka zdrowego\n i sycącego śniadania. Ich\n bogaty w składniki odżywcze\n profil i słodki smak owoców\n zapewniają energię\n na cały poranek.\n",
         " https://cutt.ly/6wZQ2fHu"],
        ["Serek wiejski z winogronami,\n orzechami włoskimi i miodem\n",
         "Ta lekka i zdrowa przekąska\n łączy serek wiejski z\n słodyczą winogron,\n chrupkością orzechów i\n miodem, tworząc idealną\n kompozycję smaków.\n",
         "https://cutt.ly/SwZQ2xdK"],
        ["Klopsiki z indyka z\n ziemniakami i zielonym groszkiem\n",
         "Soczyste klopsiki z\n indyka podane z kremowym\n sosem, młodymi ziemniakami\n i groszkiem to idealna propozycja\n na sycący i smaczny obiad.\n",
         "https://cutt.ly/JwZQ2nfj"],
        ["Smoothie z mango,\n arbuza i truskawek\n",
         "Orzeźwiające i pełne witamin,\n to smoothie to idealny sposób\n na zdrowy i szybki posiłek.\n Jego kremowa konsystencja\n i bogaty smak owoców\n zadowolą każdego.\n",
         "https://cutt.ly/iwZQ2PFx"],
        ["Gofry",
         "To danie to doskonałe\n połączenie tradycji z \n nowoczesnymi akcentami, idealne\n dla poszukiwaczy kulinarnych wrażeń.\n ",
         "https://cutt.ly/pwZQ2GLO"],
        ["Nocna owsianka z owocami",
         "To danie to perfekcyjne\n połączenie smaków i aromatów,\n które zaskoczą i zachwycą\n każdego smakosza.\n Idealne dla tych,\n którzy szukają nowych\n kulinarnych wrażeń.\n",
         "https://cutt.ly/WwZQ2Xkv"],
        ["Paluszki rybne z frytkami\n i surówką z czerwonej kapusty",
         " Klasyczne danie w\n nowoczesnej odsłonie,\n idealne dla dzieci i dorosłych.\n Chrupiące paluszki rybne\n z frytkami i świeżą surówką\n to gwarancja smacznego posiłku.\n",
         "https://cutt.ly/rwZQ21Pt"],
        ["Chlebek bananowy",
         "Słodka i aromatyczna\n przekąska, idealna\n do kawy lub herbaty.\n Ten domowy chlebek,\n pełen smaku bananów,\n zadowoli każdego łasucha.\n",
         "https://cutt.ly/pwZQ23lb"],
        ["Kotlety drobiowe z\n młodymi ziemniakami\n i sałatą",
         "Soczyste kotlety drobiowe\n podane z młodymi ziemniakami\n i świeżą sałatą to idealny\n obiad dla całej rodziny.\n Proste, ale pełne smaku.\n",
         "https://cutt.ly/SwZQ26om"],
        ["Chlebek czosnkowy",
         "Aromatyczna przekąska,\n ten domowy chlebek,\n pełen smaku czosnku,\n zadowoli każdego łasucha.\n",
         "https://cutt.ly/cwZQ9yP8"],
        ["Mocno czekoladowe trufle\n z mleka zagęszczonego i kakao",
         "To danie to perfekcyjne\n połączenie smaków i aromatów,\n które zaskoczą i zachwycą\n każdego smakosza.\n Idealne dla tych,\n którzy szukają nowych\n kulinarnych wrażeń.\n",
         " https://cutt.ly/swZQ4PUs"],
        ["Zupa fasolowa",
         "Tradycyjna, sycąca zupa\n fasolowa, pełna aromatycznych\n przypraw i warzyw.\n Doskonała na chłodne dni,\n by rozgrzać ciało i duszę.\n",
         " https://cutt.ly/cwZQ9yP8"],
        ["Tarta z burakiem,\n cebulą i kozim serem",
         " Elegancka i pełna smaku\n tarta, łącząca słodycz buraka\n z wyrazistym smakiem koziego sera.\n Idealna na wykwintną kolację\n lub imprezę.\n",
         "https://cutt.ly/XwZQ9sez"],
        ["Drożdzowe rogaliki\n z czekoladą",
         "Małe, słodkie rogaliki\n z czekoladą to idealna\n przekąska na każdą porę dnia.\n Ich miękka, drożdżowa tekstura\n i czekoladowe nadzienie sprawią,\n że nie będziesz mógł się oprzeć.\n",
         "https://cutt.ly/OwZQ9g02"],
        ["Ciasto francuskie z serem\n brie i żurawiną",
         "Wykwintne i eleganckie\n ciasto francuskie z serem brie\n i żurawiną to idealny wybór\n na wyjątkowe okazje.\n Jego delikatna tekstura i\n wyrafinowany smak zadowolą\n najbardziej wymagających gości.\n",
         " https://cutt.ly/HwZQ9zSo"]
    ]

    product_listbox8 = Listbox(frame8, font=Font(size=15))
    scrollbar8 = Scrollbar(frame8, orient="vertical", command=product_listbox8.yview)
    product_listbox8.config(yscrollcommand=scrollbar8.set, selectforeground='#06982B', selectbackground='#F3F3F3')

    product_listbox8.place(x=28, y=185, width=250, height=350)
    scrollbar8.place(x=215, y=185, height=350)

    global selected
    print(selected)


    dataframe = openpyxl.load_workbook("przepisy.xlsx")
    dataframe1 = dataframe.active

    products_potrzebne = []

    for row in range(2, dataframe1.max_row + 1):
        words = dataframe1.cell(row=row, column=3).value.split(", ")
        products_potrzebne.append([dataframe1.cell(row=row, column=2).value, words])

    for it in range(len(products)):
        # products_neccessery = products_potrzebne[it][1]
        # print(products_potrzebne[it][1], products[it][0])
        print(products_potrzebne[it][1], selected)

        if all(item in selected for item in products_potrzebne[it][1]):

        # if True:
            product_listbox8.insert("end", products[it][0]) #tutaj

            product_listbox8 = Listbox(frame8, font=Font(size=15))
            scrollbar8 = Scrollbar(frame8, orient="vertical", command=product_listbox8.yview)
            product_listbox8.config(yscrollcommand=scrollbar8.set, selectforeground='#06982B', selectbackground='#F3F3F3')

            product_listbox8.place(x=28, y=185, width=190, height=350)
            scrollbar8.place(x=215, y=185, height=350)

            # for product in products.keys():
            #     product_listbox8.insert("end", product)

            product_listbox6 = Listbox(frame8, font=Font(size=15))
            scrollbar8 = Scrollbar(frame8, orient="vertical", command=product_listbox8.yview)
            product_listbox6.config(yscrollcommand=scrollbar8.set, selectforeground='#06982B', selectbackground='#F3F3F3')

            product_listbox6.place(x=28, y=185, width=190, height=350)
            scrollbar8.place(x=215, y=185, height=350)

            for it in range(len(products)):
                if all(item in selected for item in products_potrzebne[it][1]):
                    product_listbox6.insert("end", products[it][0]) #tutaj

            product_info6 = Label(frame8, text="", font=("Inter SemiBold", 15))
            product_info6.place(x=230, y=185, width=300, height=400)

            product_listbox6.bind("<<ListboxSelect>>", update_product_info8)
            product_listbox6.selection_set(0)


            canvas.create_text(
                505.0,
                68.0,
                anchor="nw",
                text="Przepis:",
                fill="#414141",
                font=("Inter Medium", 35 * -1)
            )

            przepis = canvas.create_text(
                505.0,
                118.0,
                anchor="nw",
                text="",
                fill="#414141",
                font=("Inter Medium", 25 * -1)
            )
            link_label = Label(frame8, text="", font=("Inter Medium", 25), fg="blue", cursor="hand2")
            link_label.place(x=505.0, y=105.0)
    return frame8


# Funkcja do przełączania ramki
def switch_frame(frame_to_show, frame_to_hide):
    frame_to_hide.pack_forget()  # Ukryj nieaktywną ramkę
    frame_to_show.pack(fill='both', expand=True)


def switch_frame_5_to_8(interface5, selected_items, products):
    global selected, interface8

    # product_listbox8.delete(0, "end")
    # for product in selected:
    #     product_listbox8.insert("end", product)

    for num in selected_items:
        selected.append(products[num])

    interface8.destroy()
    interface8 = create_interface8(window)
    interface8.pack_forget()

    switch_frame(interface8, interface5)


# Główne okno aplikacji
window = Tk()
window.geometry("1000x600")
window.configure(bg="#F3F3F3")
window.title('Zdrowomix')

# Tworzenie interfejsów
# Start interfejs
interface1 = create_interface1(window)
# Menu wyboru usługi
interface2 = create_interface2(window)
# Kalkulator BMI
interface3 = create_interface3(window)
interface3.pack_forget()
# Wynik BMI
interface4 = create_interface4(window)
interface4.pack_forget()
# Spis produktow
interface5 = create_interface5(window)
interface5.pack_forget()
# Sprawdź co jesz(baba z płotem)
interface6 = create_interface6(window)
interface6.pack_forget()

interface7 = create_interface7(window)
interface7.pack_forget()
# przepisy
interface8 = create_interface8(window)
interface8.pack_forget()

# Ustawienie domyślnego interfejsu
switch_frame(interface1, interface2)
window.resizable(False, False)
window.mainloop()
